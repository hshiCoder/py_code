import openpyxl

wb = openpyxl.load_workbook('input.xlsx')
ws = wb.active

# limit为写入每个文件的数量
# start_data为Excel表数据真正开始的行，end_data为结束的行
# 当end_data为负数的时候，表示到文件的末尾
limit = 1
start_data = 2
end_data = -1

if (end_data>0 and end_data>=ws.max_row) or end_data<0:
    row_all = ws.max_row - start_data + 1
elif end_data == 0 or start_data>end_data:
    raise("希望结束的行号不能为0，或不可起始行大于结束行")
else:
    row_all = end_data - start_data + 1

row_nums = row_all
fileNameIndex = 1

def readToWrite(limit_nums):
    global fileNameIndex,start_data
    end = start_data + limit_nums
    fileName = 'output_' + str(fileNameIndex) + '.txt'
    with open(fileName,'w') as f:
        if limit>1:
            f.write('{\n')
        
        for i in range(start_data,end):
            temp1 = ws.cell(i,1).value
            temp2 = ws.cell(i,2).value
            f.write(str(temp1+temp2) + '\n')

        if limit>1:
            f.write('}')
        
    start_data = end
    fileNameIndex += 1

while row_nums>0:
    print("总数据行数：" + str(row_all) + "，等待写入数据量：" + str(row_nums) + "，正在生成文件：" + str(fileNameIndex))
    if row_nums>=limit:
        readToWrite(limit)
        row_nums -= limit
    else:
        readToWrite(row_nums)
        break

print("\n\n文件写入完成")
